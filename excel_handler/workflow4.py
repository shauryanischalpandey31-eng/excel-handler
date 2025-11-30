from __future__ import annotations

import io
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill

import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt  # noqa: E402
import numpy as np
import pandas as pd

MONTH_NAMES = [
    "January",
    "February",
    "March",
    "April",
    "May",
    "June",
    "July",
    "August",
    "September",
    "October",
    "November",
    "December",
]

MONTH_VARIANTS: Dict[str, int] = {
    "jan": 1,
    "january": 1,
    "01": 1,
    "1": 1,
    "feb": 2,
    "february": 2,
    "02": 2,
    "2": 2,
    "mar": 3,
    "march": 3,
    "03": 3,
    "3": 3,
    "apr": 4,
    "april": 4,
    "04": 4,
    "4": 4,
    "may": 5,
    "05": 5,
    "5": 5,
    "jun": 6,
    "june": 6,
    "06": 6,
    "6": 6,
    "jul": 7,
    "july": 7,
    "07": 7,
    "7": 7,
    "aug": 8,
    "august": 8,
    "08": 8,
    "8": 8,
    "sep": 9,
    "sept": 9,
    "september": 9,
    "09": 9,
    "9": 9,
    "oct": 10,
    "october": 10,
    "10": 10,
    "nov": 11,
    "november": 11,
    "11": 11,
    "dec": 12,
    "december": 12,
    "12": 12,
}

PRODUCT_KEYS = [
    "product",
    "product_name",
    "item",
    "item_name",
    "line",
    "sku",
    "code",
    "品目",
]
MONTH_KEYS = ["month", "month_name", "period", "月"]
DEMAND_KEYS = ["demand", "actual_demand", "usage", "consumption", "出荷", "需要"]
CONSUMPTION_KEYS = [
    "per_unit_consumption",
    "per unit consumption",
    "per_unit",
    "unit_consumption",
    "consumption_per_unit",
    "raw_material_per_unit",
    "unit usage",
]

DEFAULT_WINDOW = 3


@dataclass
class Workflow4Result:
    forecast_table: pd.DataFrame
    monthly_trend: pd.DataFrame
    charts: Dict[str, str]
    final_excel_path: Path
    summary: Dict[str, object]
    workflow_outputs: Dict[str, pd.DataFrame]


def run_workflow4_pipeline(
    uploaded_file,
    *,
    processed_dir: Path,
    charts_dir: Path,
    original_file_path: Optional[Path] = None,
) -> Workflow4Result:
    """
    High-level orchestration that loads the uploaded workbook, parses prior
    workflow outputs, performs forecasting, saves the consolidated Excel file,
    and generates the requested charts.
    
    Args:
        uploaded_file: File-like object containing the Excel file or file path
        processed_dir: Directory to save processed files
        charts_dir: Directory to save generated charts
        original_file_path: Optional path to original file for writing back results
    """
    # Handle both file-like objects and file paths
    if isinstance(uploaded_file, (str, Path)):
        file_path = Path(uploaded_file)
        workbook = pd.ExcelFile(file_path)
    else:
        file_bytes = uploaded_file.read()
        uploaded_file.seek(0)
        workbook = pd.ExcelFile(io.BytesIO(file_bytes))
    
    base_sheet = workbook.parse(workbook.sheet_names[0])
    workflow_outputs = _collect_workflow_outputs(workbook)
    normalized = _normalize_input_dataframe(base_sheet)
    forecast_table, demand_trend, method_map = _build_forecast_tables(normalized)
    workflow_outputs["Workflow 4"] = forecast_table
    final_excel_path = _write_final_excel(workflow_outputs, processed_dir, original_file_path)
    chart_paths = _generate_charts(demand_trend, forecast_table, charts_dir)
    summary = {
        "products": len(forecast_table),
        "total_forecast": float(forecast_table["Forecast Demand"].sum()),
        "total_raw_material": float(forecast_table["Raw Material Needed"].sum()),
        "methods": method_map,
    }
    return Workflow4Result(
        forecast_table=forecast_table,
        monthly_trend=demand_trend,
        charts=chart_paths,
        final_excel_path=final_excel_path,
        summary=summary,
        workflow_outputs=workflow_outputs,
    )


def _collect_workflow_outputs(workbook: pd.ExcelFile) -> Dict[str, pd.DataFrame]:
    outputs: Dict[str, pd.DataFrame] = {}
    for sheet in workbook.sheet_names:
        if "workflow" in sheet.lower():
            outputs[sheet] = workbook.parse(sheet)
    for idx in range(1, 4):
        label = f"Workflow {idx}"
        outputs.setdefault(label, pd.DataFrame({"Info": ["No data provided"]}))
    return outputs


def _normalize_input_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    frame = df.copy()
    frame.columns = [str(col).strip() for col in frame.columns]
    product_col = _find_column(frame, PRODUCT_KEYS) or frame.columns[0]
    month_col = _find_column(frame, MONTH_KEYS)
    demand_col = _find_column(frame, DEMAND_KEYS)
    consumption_col = _find_column(frame, CONSUMPTION_KEYS)
    if month_col and demand_col and consumption_col:
        rename_map = {
            product_col: "product",
            month_col: "month",
            demand_col: "demand",
            consumption_col: "per_unit_consumption",
        }
        normalized = (
            frame.rename(columns=rename_map)[
                ["product", "month", "demand", "per_unit_consumption"]
            ]
            .dropna(subset=["product", "month", "demand", "per_unit_consumption"])
            .copy()
        )
        normalized["product"] = normalized["product"].astype(str).str.strip()
        normalized["month"] = normalized["month"].apply(
            lambda value: _canonical_month(value) or value
        )
        normalized["demand"] = pd.to_numeric(normalized["demand"], errors="coerce")
        normalized["per_unit_consumption"] = pd.to_numeric(
            normalized["per_unit_consumption"], errors="coerce"
        )
        normalized = normalized.dropna(
            subset=["demand", "per_unit_consumption", "month"]
        )
        return normalized
    if consumption_col is None:
        raise ValueError(
            "Unable to locate the per-unit consumption column. Please ensure the "
            "uploaded file retains the column produced by workflow 3."
        )
    month_columns = _detect_month_columns(frame)
    if not month_columns:
        raise ValueError(
            "Could not detect any monthly columns. Ensure headers or the first few "
            "rows include month names (e.g., April, 11月, etc.)."
        )
    records: List[Dict[str, object]] = []
    for _, row in frame.iterrows():
        product_value = row.get(product_col)
        if pd.isna(product_value):
            continue
        product_text = str(product_value).strip()
        if not product_text or _canonical_month(product_text):
            continue  # skip rows that only hold month labels
        per_unit_value = row.get(consumption_col)
        for column, canonical_month in month_columns.items():
            cell_value = row.get(column)
            if pd.isna(cell_value):
                continue
            if isinstance(cell_value, str) and _canonical_month(cell_value):
                continue
            numeric_value = _coerce_number(cell_value)
            if numeric_value is None:
                continue
            records.append(
                {
                    "product": product_text,
                    "month": canonical_month,
                    "demand": numeric_value,
                    "per_unit_consumption": per_unit_value,
                }
            )
    long_df = pd.DataFrame(records)
    if long_df.empty:
        raise ValueError(
            "No monthly demand values could be parsed from the supplied sheet."
        )
    long_df["per_unit_consumption"] = pd.to_numeric(
        long_df["per_unit_consumption"], errors="coerce"
    )
    long_df["per_unit_consumption"] = long_df.groupby("product")[
        "per_unit_consumption"
    ].ffill().bfill()
    if long_df["per_unit_consumption"].isna().any():
        raise ValueError(
            "Per-unit consumption values are missing for some products. "
            "Please ensure workflow 3 outputs are included."
        )
    return long_df.dropna(subset=["demand"])


def _detect_month_columns(df: pd.DataFrame) -> Dict[str, str]:
    mapping: Dict[str, str] = {}
    for column in df.columns:
        normalized = _canonical_month(column)
        if normalized:
            mapping[column] = normalized
            continue
        sample_values = (
            df[column].dropna().astype(str).str.strip().head(4).tolist()
        )
        for sample in sample_values:
            normalized_value = _canonical_month(sample)
            if normalized_value:
                mapping[column] = normalized_value
                break
    return mapping


def _build_forecast_tables(
    df: pd.DataFrame,
    *,
    window: int = DEFAULT_WINDOW,
) -> Tuple[pd.DataFrame, pd.DataFrame, Dict[str, str]]:
    """
    Build forecast tables using 3-month moving average for next-month prediction.
    Implements proper stock usage forecasting with business logic.
    """
    records: List[Dict[str, object]] = []
    trend_frames: List[pd.DataFrame] = []
    methods: Dict[str, str] = {}
    
    for product, group in df.groupby("product"):
        cleaned = (
            group.dropna(subset=["demand"])
            .copy()
            .assign(month_index=lambda g: g["month"].apply(_month_index))
            .dropna(subset=["month_index"])
            .sort_values("month_index")
        )
        if cleaned.empty:
            continue
        
        trend_frames.append(
            cleaned[["month", "demand"]].assign(product=product).reset_index(drop=True)
        )
        
        demand_series = cleaned["demand"].values
        month_indices = cleaned["month_index"].values
        
        # Calculate next-month forecast using 3-month moving average
        if len(demand_series) >= window:
            # Use the last 3 months for moving average
            forecast_value = float(np.mean(demand_series[-window:]))
            method = f"{window}-month moving average"
        elif len(demand_series) >= 2:
            # If less than 3 months, use available data
            forecast_value = float(np.mean(demand_series))
            method = f"{len(demand_series)}-month average"
        else:
            # Single data point or insufficient data
            forecast_value = float(demand_series[-1]) if len(demand_series) > 0 else 0.0
            method = "single period or insufficient data"
        
        # Ensure forecast is non-negative
        forecast_value = max(0.0, forecast_value)
        
        per_unit_series = cleaned["per_unit_consumption"].dropna()
        if per_unit_series.empty:
            raise ValueError(
                f"Missing per-unit consumption values for product '{product}'."
            )
        per_unit_value = float(per_unit_series.iloc[-1])
        
        # Calculate raw material needed
        raw_material_needed = forecast_value * per_unit_value
        
        records.append(
            {
                "Product": product,
                "Forecast Demand": round(forecast_value, 2),
                "Per Unit Consumption": round(per_unit_value, 4),
                "Raw Material Needed": round(raw_material_needed, 2),
            }
        )
        methods[product] = method
    
    forecast_table = pd.DataFrame(records)
    if forecast_table.empty:
        raise ValueError("Forecast table is empty. Please provide historical demand.")
    
    demand_trend = (
        pd.concat(trend_frames, ignore_index=True)
        if trend_frames
        else pd.DataFrame(columns=["month", "demand", "product"])
    )
    if not demand_trend.empty:
        demand_trend["month_index"] = demand_trend["month"].apply(_month_index)
        demand_trend = demand_trend.dropna(subset=["month_index"]).sort_values(
            "month_index"
        )
    
    return forecast_table, demand_trend.reset_index(drop=True), methods


def _write_final_excel(
    workflow_outputs: Dict[str, pd.DataFrame],
    processed_dir: Path,
    original_workbook_path: Optional[Path] = None,
) -> Path:
    """
    Write final Excel file with proper formatting.
    If original_workbook_path is provided, writes results back to original structure.
    """
    from openpyxl.styles import Border, Side
    from openpyxl.utils import get_column_letter
    
    processed_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y_%m_%d_%H%M%S")
    target_path = processed_dir / f"final_output_{timestamp}.xlsx"
    
    # If original workbook path is provided, load it and write back
    if original_workbook_path and original_workbook_path.exists():
        try:
            wb = load_workbook(original_workbook_path)
            # Add or update Workflow 4 sheet
            if "Workflow 4" in wb.sheetnames:
                ws = wb["Workflow 4"]
                wb.remove(ws)
            ws = wb.create_sheet("Workflow 4")
            
            # Write Workflow 4 data with formatting
            if "Workflow 4" in workflow_outputs:
                df = workflow_outputs["Workflow 4"]
                if not df.empty:
                    # Write headers
                    for col_idx, col_name in enumerate(df.columns, start=1):
                        cell = ws.cell(row=1, column=col_idx, value=col_name)
                        cell.font = Font(bold=True, size=11)
                        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                        cell.font = Font(bold=True, size=11, color="FFFFFF")
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    
                    # Write data
                    for row_idx, (_, row) in enumerate(df.iterrows(), start=2):
                        for col_idx, value in enumerate(row, start=1):
                            cell = ws.cell(row=row_idx, column=col_idx, value=value)
                            cell.alignment = Alignment(horizontal="right", vertical="center")
                    
                    # Auto-adjust column widths
                    for col_idx in range(1, len(df.columns) + 1):
                        col_letter = get_column_letter(col_idx)
                        max_length = 0
                        for row in ws[col_letter]:
                            try:
                                if len(str(row.value)) > max_length:
                                    max_length = len(str(row.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 50)
                        ws.column_dimensions[col_letter].width = adjusted_width
            
            wb.save(target_path)
            wb.close()
            return target_path
        except Exception as e:
            # Fall back to pandas ExcelWriter if openpyxl operations fail
            pass
    
    # Default: Use pandas ExcelWriter with formatting
    with pd.ExcelWriter(target_path, engine="openpyxl") as writer:
        for sheet_name, data in workflow_outputs.items():
            sheet_df = data if not data.empty else pd.DataFrame({"Info": ["No data"]})
            sheet_df.to_excel(writer, sheet_name=sheet_name[:31], index=False)
            
            # Apply formatting using openpyxl
            workbook = writer.book
            worksheet = writer.sheets[sheet_name[:31]]
            
            # Format header row
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, size=11, color="FFFFFF")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
            
            # Format data cells
            data_alignment = Alignment(horizontal="right", vertical="center")
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                for cell in row:
                    cell.alignment = data_alignment
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
    
    return target_path


def write_results_to_original_excel(
    original_file_path: Path,
    forecast_table: pd.DataFrame,
    output_path: Path,
) -> Path:
    """
    Write Workflow-4 forecast results back into the original Excel file structure.
    Preserves all original sheets, formatting, and layout.
    """
    try:
        wb = load_workbook(original_file_path)
        
        # Add or update Workflow 4 sheet
        if "Workflow 4" in wb.sheetnames:
            wb.remove(wb["Workflow 4"])
        ws = wb.create_sheet("Workflow 4")
        
        # Write headers with formatting
        headers = ["Product", "Forecast Demand", "Per Unit Consumption", "Raw Material Needed"]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True, size=11, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Write data rows
        for row_idx, (_, row) in enumerate(forecast_table.iterrows(), start=2):
            ws.cell(row=row_idx, column=1, value=row["Product"])
            ws.cell(row=row_idx, column=2, value=row["Forecast Demand"])
            ws.cell(row=row_idx, column=3, value=row["Per Unit Consumption"])
            ws.cell(row=row_idx, column=4, value=row["Raw Material Needed"])
            
            # Format data cells
            for col_idx in range(1, 5):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.alignment = Alignment(horizontal="right" if col_idx > 1 else "left", vertical="center")
        
        # Auto-adjust column widths
        from openpyxl.utils import get_column_letter
        for col_idx in range(1, 5):
            col_letter = get_column_letter(col_idx)
            max_length = len(headers[col_idx - 1])
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_length + 2, 50)
        
        wb.save(output_path)
        wb.close()
        return output_path
    except Exception as e:
        # If writing to original fails, create a new file
        raise ValueError(f"Failed to write to original Excel structure: {str(e)}")


def _generate_charts(
    demand_trend: pd.DataFrame,
    raw_material_table: pd.DataFrame,
    charts_dir: Path,
) -> Dict[str, str]:
    charts_dir.mkdir(parents=True, exist_ok=True)
    demand_chart_path = charts_dir / "demand_plot.png"
    raw_chart_path = charts_dir / "raw_material_plot.png"
    if not demand_trend.empty:
        agg_trend = (
            demand_trend.groupby("month", as_index=False)["demand"]
            .sum()
            .assign(month_index=lambda df: df["month"].apply(_month_index))
            .dropna(subset=["month_index"])
            .sort_values("month_index")
        )
        plt.figure(figsize=(10, 4))
        plt.plot(agg_trend["month"], agg_trend["demand"], marker="o", linewidth=2)
        plt.title("Monthly Demand Trend")
        plt.xlabel("Month")
        plt.ylabel("Units")
        plt.grid(True, linestyle="--", alpha=0.3)
        plt.tight_layout()
        plt.savefig(demand_chart_path, dpi=150)
        plt.close()
    else:
        plt.figure(figsize=(10, 4))
        plt.text(0.5, 0.5, "No demand data", ha="center", va="center")
        plt.axis("off")
        plt.savefig(demand_chart_path, dpi=150)
        plt.close()
    plt.figure(figsize=(8, 4))
    plt.bar(
        raw_material_table["Product"],
        raw_material_table["Raw Material Needed"],
        color="#4e79a7",
    )
    plt.title("Raw Material Requirement by Product")
    plt.xlabel("Product")
    plt.ylabel("Raw Material Needed")
    plt.xticks(rotation=45, ha="right")
    plt.tight_layout()
    plt.savefig(raw_chart_path, dpi=150)
    plt.close()
    return {
        "demand": f"charts/{demand_chart_path.name}",
        "raw_material": f"charts/{raw_chart_path.name}",
    }


def _find_column(df: pd.DataFrame, candidates: Iterable[str]) -> Optional[str]:
    lowered = {candidate.lower(): candidate for candidate in candidates}
    for column in df.columns:
        col_value = str(column).strip()
        col_lower = col_value.lower()
        if col_lower in lowered:
            return column
        for candidate in candidates:
            if candidate.lower() in col_lower:
                return column
    return None


def _canonical_month(value) -> Optional[str]:
    if value is None or (isinstance(value, float) and np.isnan(value)):
        return None
    text = str(value).strip().lower()
    if not text:
        return None
    if text.endswith("月"):
        text = text.replace("月", "")
    text = text.replace(".", "")
    if text in MONTH_VARIANTS:
        idx = MONTH_VARIANTS[text]
        return MONTH_NAMES[idx - 1]
    if text.isdigit():
        idx = int(text)
        if 1 <= idx <= 12:
            return MONTH_NAMES[idx - 1]
    return None


def _month_index(value) -> Optional[int]:
    canonical = _canonical_month(value)
    if canonical is None:
        return None
    return MONTH_NAMES.index(canonical) + 1


def _coerce_number(value) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float, np.number)):
        return float(value)
    try:
        return float(str(value).replace(",", ""))
    except ValueError:
        return None

